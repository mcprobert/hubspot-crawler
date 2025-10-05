"""
Tests for Excel (.xlsx) output format.

Covers:
- Excel file creation
- Column headers
- Data types (booleans, numbers, strings)
- Fields with commas and special characters
- openpyxl import handling
"""

import pytest
import asyncio
import tempfile
import os
from hubspot_crawler.detector import make_result
from hubspot_crawler.crawler import flatten_result_for_csv


class TestExcelOutput:
    """Test Excel file generation"""

    @pytest.mark.asyncio
    async def test_excel_file_creation(self):
        """Should create a valid Excel file with headers"""
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed")

        from hubspot_crawler.crawler import excel_writer_worker

        # Create temp file
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = tmp.name

        try:
            # Create queue and add test data
            queue = asyncio.Queue()
            evidence = []
            result = make_result("https://example.com", "https://example.com", evidence)
            await queue.put(result)
            await queue.put(None)  # Poison pill

            # Run writer
            await excel_writer_worker(queue, tmp_path)

            # Verify file was created
            assert os.path.exists(tmp_path)
            assert os.path.getsize(tmp_path) > 0

            # Verify file can be opened and has correct structure
            wb = openpyxl.load_workbook(tmp_path)
            ws = wb.active

            # Check headers (row 1)
            expected_headers = [
                "original_url", "final_url", "timestamp", "hubspot_detected", "tracking", "cms_hosting", "confidence",
                "forms", "chat", "ctas_legacy", "meetings", "video", "email_tracking",
                "hub_ids", "hub_id_count", "evidence_count", "http_status", "page_title", "page_description"
            ]

            headers = [cell.value for cell in ws[1]]
            assert headers == expected_headers

            # Check header row is bold
            for cell in ws[1]:
                assert cell.font.bold is True

            # Check data row exists (row 2)
            assert ws[2][0].value == "https://example.com"  # original_url
            assert ws[2][1].value == "https://example.com"  # final_url

            wb.close()
        finally:
            # Cleanup
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)

    @pytest.mark.asyncio
    async def test_excel_with_commas_in_fields(self):
        """Should handle commas in text fields without issues"""
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed")

        from hubspot_crawler.crawler import excel_writer_worker

        # Create temp file
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = tmp.name

        try:
            # Create queue and add test data with commas
            queue = asyncio.Queue()
            evidence = []
            result = make_result("https://example.com", "https://example.com", evidence)

            # Add page metadata with commas
            result["page_metadata"] = {
                "title": "Marketing, Sales, and Service Software",
                "description": "A platform for marketing, sales, service, and more"
            }

            await queue.put(result)
            await queue.put(None)  # Poison pill

            # Run writer
            await excel_writer_worker(queue, tmp_path)

            # Verify file can be read and commas are preserved
            wb = openpyxl.load_workbook(tmp_path)
            ws = wb.active

            # Check that commas are preserved in title and description
            page_title_col = 17  # 0-indexed position in fieldnames
            page_desc_col = 18

            title_value = ws.cell(row=2, column=page_title_col + 1).value
            desc_value = ws.cell(row=2, column=page_desc_col + 1).value

            assert "Marketing, Sales, and Service Software" == title_value
            assert "A platform for marketing, sales, service, and more" == desc_value

            wb.close()
        finally:
            # Cleanup
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)

    @pytest.mark.asyncio
    async def test_excel_boolean_data_types(self):
        """Should preserve boolean data types in Excel"""
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed")

        from hubspot_crawler.crawler import excel_writer_worker

        # Create temp file
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = tmp.name

        try:
            # Create queue and add test data with HubSpot detection
            queue = asyncio.Queue()
            evidence = [{
                "category": "tracking",
                "patternId": "tracking_loader_script",
                "match": "js.hs-scripts.com/123.js",
                "source": "html",
                "hubId": 123,
                "confidence": "definitive"
            }]
            result = make_result("https://example.com", "https://example.com", evidence)

            await queue.put(result)
            await queue.put(None)  # Poison pill

            # Run writer
            await excel_writer_worker(queue, tmp_path)

            # Verify boolean values
            wb = openpyxl.load_workbook(tmp_path)
            ws = wb.active

            # hubspot_detected should be True
            hubspot_detected_col = 4  # 0-indexed position
            assert ws.cell(row=2, column=hubspot_detected_col).value is True

            # tracking should be True
            tracking_col = 5
            assert ws.cell(row=2, column=tracking_col).value is True

            # cms_hosting should be False
            cms_hosting_col = 6
            assert ws.cell(row=2, column=cms_hosting_col).value is False

            wb.close()
        finally:
            # Cleanup
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)

    @pytest.mark.asyncio
    async def test_excel_multiple_rows(self):
        """Should handle multiple result rows"""
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed")

        from hubspot_crawler.crawler import excel_writer_worker

        # Create temp file
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = tmp.name

        try:
            # Create queue and add multiple results
            queue = asyncio.Queue()

            for i in range(5):
                evidence = []
                result = make_result(f"https://example{i}.com", f"https://example{i}.com", evidence)
                await queue.put(result)

            await queue.put(None)  # Poison pill

            # Run writer
            await excel_writer_worker(queue, tmp_path)

            # Verify all rows were written
            wb = openpyxl.load_workbook(tmp_path)
            ws = wb.active

            # Should have header + 5 data rows
            assert ws.max_row == 6

            # Check URLs are correct
            for i in range(5):
                assert ws.cell(row=i + 2, column=1).value == f"https://example{i}.com"

            wb.close()
        finally:
            # Cleanup
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)

    def test_excel_import_error_handling(self):
        """Should raise helpful error if openpyxl not installed"""
        # Temporarily hide openpyxl if it exists
        import sys
        old_openpyxl = sys.modules.get('openpyxl')
        if 'openpyxl' in sys.modules:
            del sys.modules['openpyxl']

        try:
            # This should raise RuntimeError with install instructions
            from hubspot_crawler.crawler import excel_writer_worker
            import asyncio

            queue = asyncio.Queue()

            # Try to create writer (should fail with import error)
            async def test_import():
                # Mock the import to fail
                import builtins
                real_import = builtins.__import__

                def mock_import(name, *args, **kwargs):
                    if name == 'openpyxl':
                        raise ImportError("No module named 'openpyxl'")
                    return real_import(name, *args, **kwargs)

                builtins.__import__ = mock_import
                try:
                    with pytest.raises(RuntimeError, match="openpyxl not installed"):
                        await excel_writer_worker(queue, "test.xlsx")
                finally:
                    builtins.__import__ = real_import

            asyncio.run(test_import())

        finally:
            # Restore openpyxl
            if old_openpyxl:
                sys.modules['openpyxl'] = old_openpyxl
